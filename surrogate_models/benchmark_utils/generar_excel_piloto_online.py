"""
Genera Excel de comparación de configuraciones del piloto online.
Para cada algoritmo (AGE, DE, SHADE) rankea todas sus configuraciones
por cec_error medio en cada función (TACOLAB-style), promedia el rango
sobre las 7 funciones y ordena por rank medio.

Estructura:
  - Hojas AGE/DE/SHADE: p | cd | rt | Rank medio | Pos. | Error f3..f27
  - Hoja "Resumen global": p | cd | rt | Pos.AGE | Pos.DE | Pos.SHADE | Peor pos.
    Ordenado por Peor pos. (minimax). Verde en Peor pos.
"""
from __future__ import annotations

import glob
import re
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

PILOT_GLOB = "results/cec/piloto_online_grid_*/f*/runs.csv"
OUT_PATH = Path("memoria/tablas/piloto_online_comparativa.xlsx")

ALGO_MAP = {"age": "AGE", "de": "DE", "shade": "SHADE"}

GREEN = PatternFill("solid", fgColor="C6EFCE")


def parse_params(adaptacion: str) -> tuple[float, int, float] | None:
    m = re.match(r"[a-z]{2}_p(\d+)_c(\d+)_rt(\d+)", adaptacion)
    if not m:
        return None
    return int(m.group(1)) / 100, int(m.group(2)), int(m.group(3)) / 100


def load_pilot() -> pd.DataFrame:
    dfs = [pd.read_csv(f) for f in sorted(glob.glob(PILOT_GLOB))]
    df = pd.concat(dfs, ignore_index=True)
    parsed = df["adaptacion"].map(parse_params)
    df["p"] = parsed.map(lambda x: x[0] if x else None)
    df["cd"] = parsed.map(lambda x: x[1] if x else None)
    df["rt"] = parsed.map(lambda x: x[2] if x else None)
    df = df[df["p"].notna()].copy()
    df = df[(df["cd"] > 0) & (df["p"] > 0)]
    return df


def rank_configs(df_algo: pd.DataFrame) -> pd.DataFrame:
    mean_err = (
        df_algo.groupby(["p", "cd", "rt", "cec_funcid"])["cec_error"]
        .mean()
        .reset_index()
    )
    mean_err["rank"] = mean_err.groupby("cec_funcid")["cec_error"].rank(
        ascending=True, method="average"
    )
    rank_medio = mean_err.groupby(["p", "cd", "rt"])["rank"].mean().rename("rank_medio")
    err_pivot = mean_err.pivot_table(
        index=["p", "cd", "rt"], columns="cec_funcid", values="cec_error"
    )
    err_pivot.columns = [f"f{int(c)}" for c in err_pivot.columns]
    result = rank_medio.to_frame().join(err_pivot).reset_index()
    result = result.sort_values("rank_medio").reset_index(drop=True)
    result["pos"] = range(1, len(result) + 1)
    return result


def _bold_header(ws, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)


def _green_best(ws, data_rows: list[int], col: int) -> None:
    vals = [ws.cell(row=r, column=col).value for r in data_rows]
    vals_num = [v for v in vals if isinstance(v, (int, float))]
    if not vals_num:
        return
    best = min(vals_num)
    for r in data_rows:
        v = ws.cell(row=r, column=col).value
        if isinstance(v, (int, float)) and abs(v - best) < 1e-9:
            ws.cell(row=r, column=col).fill = GREEN


def _autowidth(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 18)


def write_algo_sheet(wb: openpyxl.Workbook, algo: str, ranked: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=ALGO_MAP[algo])
    headers = ["p", "cd", "rt", "Rank medio", "Pos."]
    ws.append(headers)
    _bold_header(ws, len(headers))

    data_rows = []
    for i, row in enumerate(ranked.itertuples(index=False), start=2):
        ws.append([row.p, int(row.cd), row.rt, round(row.rank_medio, 3), row.pos])
        data_rows.append(i)

    _green_best(ws, data_rows, col=4)  # Rank medio
    _autowidth(ws)


def write_summary_sheet(wb: openpyxl.Workbook, ranked_by_algo: dict[str, pd.DataFrame]) -> None:
    # Construir mapa (p,cd,rt) -> pos por algo
    all_configs: dict[tuple, dict] = {}
    for algo, ranked in ranked_by_algo.items():
        for row in ranked.itertuples(index=False):
            key = (row.p, int(row.cd), row.rt)
            if key not in all_configs:
                all_configs[key] = {}
            all_configs[key][f"pos_{algo}"] = row.pos

    def worst_pos(entry: dict) -> int:
        vals = [entry[f"pos_{a}"] for a in ["age", "de", "shade"] if f"pos_{a}" in entry]
        return max(vals) if vals else 999

    summary = sorted(all_configs.items(), key=lambda x: worst_pos(x[1]))

    ws = wb.create_sheet(title="Resumen global", index=0)
    headers = ["p", "cd", "rt", "Pos. AGE", "Pos. DE", "Pos. SHADE", "Peor pos."]
    ws.append(headers)
    _bold_header(ws, len(headers))

    data_rows = []
    for i, (key, entry) in enumerate(summary, start=2):
        ws.append([
            key[0], key[1], key[2],
            entry.get("pos_age", ""),
            entry.get("pos_de", ""),
            entry.get("pos_shade", ""),
            worst_pos(entry),
        ])
        data_rows.append(i)

    _green_best(ws, data_rows, col=7)  # Peor pos.
    _autowidth(ws)


def main() -> None:
    print("Cargando datos del piloto...")
    df = load_pilot()
    print(f"  {len(df)} filas | algoritmos: {sorted(df['algoritmo'].unique())}")

    ranked_by_algo = {}
    for algo in ["age", "de", "shade"]:
        sub = df[df["algoritmo"] == algo]
        ranked_by_algo[algo] = rank_configs(sub)
        print(f"  {ALGO_MAP[algo]}: {len(ranked_by_algo[algo])} configuraciones")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb, ranked_by_algo)
    for algo in ["age", "de", "shade"]:
        write_algo_sheet(wb, algo, ranked_by_algo[algo])

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Guardado en {OUT_PATH}")


if __name__ == "__main__":
    main()
