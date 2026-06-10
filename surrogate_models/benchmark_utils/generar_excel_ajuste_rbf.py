"""
Genera Excel de ajuste de hiperparámetros RBF.
- Hoja "Resumen global": Kernel | ε | Smoothing | Vecinos | Pos. AGE | Pos. DE | Pos. SHADE | Peor pos.
  Ordenado por Peor pos. (minimax). Verde en Peor pos.
- Hojas "AGE", "DE", "SHADE": Kernel | ε | Smoothing | Vecinos | Rank medio | Pos.
  Ordenadas por Rank medio. Verde en Rank medio.
"""
from __future__ import annotations

import ast
import csv
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(
    "results/cec/cec2017_d10_tam50_reinicio_seleccionado/"
    "benchmarking_surrogates_hyperparameter_selection/future_next/rbf/no_acumulativo"
)
OUT_PATH = Path("memoria/tablas/ajuste_interno_rbf_por_algoritmo.xlsx")

ALGORITHMS = ["age", "de", "shade"]
ALGO_LABELS = {"age": "AGE", "de": "DE", "shade": "SHADE"}

GREEN = PatternFill("solid", fgColor="C6EFCE")


def _clean_repr(raw: str) -> str:
    s = re.sub(r"np\.float64\(([^)]+)\)", r"\1", raw)
    s = re.sub(r"np\.int64\(([^)]+)\)", r"\1", s)
    s = re.sub(r"np\.nan\b", "None", s)
    return s


def _params_key(p: dict) -> tuple:
    return (
        p.get("kernel", ""),
        float(p.get("epsilon", 0)),
        float(p.get("smoothing", 0)),
        int(p.get("neighbors", 0)),
    )


def _rank_configs(configs: list[dict]) -> list[float]:
    scores = [
        c["score"] if c.get("error") is None and c.get("score") is not None else float("-inf")
        for c in configs
    ]
    sorted_unique = sorted(set(scores), reverse=True)
    rank_map: dict[float, float] = {}
    pos = 1
    for val in sorted_unique:
        count = scores.count(val)
        rank_map[val] = pos + (count - 1) / 2.0
        pos += count
    return [rank_map[s] for s in scores]


def load_data() -> dict[str, dict[tuple, dict]]:
    data: dict[str, dict[tuple, dict]] = {
        algo: defaultdict(lambda: {"ranks": [], "params": None})
        for algo in ALGORITHMS
    }
    for func_dir in sorted(BASE_DIR.iterdir()):
        if not func_dir.is_dir() or not func_dir.name.startswith("f"):
            continue
        for algo in ALGORITHMS:
            runs_path = func_dir / algo / "rbf" / "rbf_runs.csv"
            if not runs_path.exists():
                continue
            rows = list(csv.DictReader(runs_path.read_text(encoding="utf-8").splitlines()))
            for row in rows:
                raw = row.get("tuning_resultados", "")
                if not raw:
                    continue
                try:
                    configs = ast.literal_eval(_clean_repr(raw))
                except Exception:
                    continue
                if not configs:
                    continue
                ranks = _rank_configs(configs)
                for cfg, rank in zip(configs, ranks):
                    if cfg.get("error") is not None or cfg.get("score") is None:
                        continue
                    key = _params_key(cfg["params"])
                    entry = data[algo][key]
                    if entry["params"] is None:
                        entry["params"] = cfg["params"]
                    entry["ranks"].append(rank)
    return data


def aggregate(data: dict[str, dict[tuple, dict]]) -> dict[str, list[dict]]:
    result = {}
    for algo, configs in data.items():
        rows = []
        for key, entry in configs.items():
            if not entry["ranks"] or entry["params"] is None:
                continue
            rows.append({
                "key": key,
                "params": entry["params"],
                "rank": round(sum(entry["ranks"]) / len(entry["ranks"]), 3),
            })
        rows.sort(key=lambda r: r["rank"])
        for pos, r in enumerate(rows, start=1):
            r["pos"] = pos
        result[algo] = rows
    return result


def _green_best(ws, data_rows: list, col: int) -> None:
    vals = [ws.cell(row=r, column=col).value for r in data_rows]
    vals_num = [v for v in vals if isinstance(v, (int, float))]
    if not vals_num:
        return
    best = min(vals_num)
    for r in data_rows:
        v = ws.cell(row=r, column=col).value
        if isinstance(v, (int, float)) and abs(v - best) < 1e-9:
            ws.cell(row=r, column=col).fill = GREEN


def _bold_header(ws, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)


def _autowidth(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)


def write_algo_sheet(wb: openpyxl.Workbook, algo: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(title=ALGO_LABELS[algo])
    headers = ["Kernel", "ε", "Smoothing", "Vecinos", "Rank medio", "Pos."]
    ws.append(headers)
    _bold_header(ws, len(headers))

    data_rows = []
    for i, r in enumerate(rows, start=2):
        p = r["params"]
        ws.append([
            p.get("kernel", ""),
            float(p.get("epsilon", 0)),
            float(p.get("smoothing", 0)),
            int(p.get("neighbors", 0)),
            r["rank"],
            r["pos"],
        ])
        data_rows.append(i)

    _green_best(ws, data_rows, col=5)  # Rank medio
    _autowidth(ws)


def write_summary_sheet(wb: openpyxl.Workbook, agg: dict[str, list[dict]]) -> None:
    # Construir mapa key -> pos por algo
    pos_map: dict[tuple, dict] = {}
    for algo, rows in agg.items():
        for r in rows:
            k = r["key"]
            if k not in pos_map:
                pos_map[k] = {"params": r["params"]}
            pos_map[k][f"pos_{algo}"] = r["pos"]

    def worst_pos(entry: dict) -> int:
        vals = [entry[f"pos_{a}"] for a in ALGORITHMS if f"pos_{a}" in entry]
        return max(vals) if vals else 999

    summary = sorted(pos_map.values(), key=worst_pos)

    ws = wb.create_sheet(title="Resumen global", index=0)
    headers = ["Kernel", "ε", "Smoothing", "Vecinos", "Pos. AGE", "Pos. DE", "Pos. SHADE", "Peor pos."]
    ws.append(headers)
    _bold_header(ws, len(headers))

    data_rows = []
    for i, entry in enumerate(summary, start=2):
        p = entry["params"]
        ws.append([
            p.get("kernel", ""),
            float(p.get("epsilon", 0)),
            float(p.get("smoothing", 0)),
            int(p.get("neighbors", 0)),
            entry.get("pos_age", ""),
            entry.get("pos_de", ""),
            entry.get("pos_shade", ""),
            worst_pos(entry),
        ])
        data_rows.append(i)

    _green_best(ws, data_rows, col=8)  # Peor pos.
    _autowidth(ws)


def main() -> None:
    print(f"Cargando datos desde {BASE_DIR} ...")
    raw = load_data()
    agg = aggregate(raw)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb, agg)
    for algo in ALGORITHMS:
        write_algo_sheet(wb, algo, agg[algo])

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Guardado en {OUT_PATH}")


if __name__ == "__main__":
    main()
